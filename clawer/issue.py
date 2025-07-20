import os
import time
import json
import pandas as pd
import re
from collections import defaultdict
from dotenv import load_dotenv
from github import Github, Auth, GithubException, RateLimitExceededException
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from log_config import logger


# 设置基本的日志输出
class SimpleLogger:
    def info(self, msg):
        print(f"[INFO] {msg}")

    def warning(self, msg):
        print(f"[WARNING] {msg}")

    def error(self, msg):
        print(f"[ERROR] {msg}")


def write_row_to_excel(_ws, _row_data, _row_num):
    """
    将一行数据写入Excel工作表中的指定行号，如果遇到非法字符错误则跳过该行。
    """
    try:
        for col_num, value in enumerate(_row_data, start=1):
            _ws.cell(row=_row_num, column=col_num, value=value)
    except IllegalCharacterError:
        logger.warning(f"在写入第 {_row_num} 行时遇到非法字符，已跳过。")
        return False
    return True


def save_progress(repo_name, last_issue_number):
    """保存处理进度到文件"""
    progress_file = f'../issue_results/{repo_name.replace("/", "_")}_progress.json'
    with open(progress_file, 'w') as f:
        json.dump({"last_issue_number": last_issue_number}, f)


def load_progress(repo_name):
    """加载保存的进度"""
    progress_file = f'../issue_results/{repo_name.replace("/", "_")}_progress.json'
    if os.path.exists(progress_file):
        with open(progress_file, 'r') as f:
            data = json.load(f)
            return data.get("last_issue_number", 0)
    return 0


def analyze_images_in_content(content):
    """
    分析内容中的图片链接
    """
    if not content:
        return {'markdown_images': [], 'html_images': []}

    # 匹配markdown图片语法: ![alt](url)
    image_pattern = r'!\[([^\]]*)\]\(([^)]+)\)'
    images = re.findall(image_pattern, content)

    # 匹配HTML img标签
    html_img_pattern = r'<img[^>]*src="([^"]*)"[^>]*>'
    html_images = re.findall(html_img_pattern, content)


    return {
        'markdown_images': images,  # [(alt_text, url), ...]
        'html_images': html_images,  # [url, ...]
    }


def get_issue_to_excel_with_analysis(repo_name):
    """
    使用issue_2的逻辑爬取issues到Excel，并同时进行图片和标签分析
    """
    load_dotenv()
    my_github_token = os.getenv("MY_GITHUB_TOKEN")
    if not my_github_token:
        raise ValueError("You need to set a GITHUB_TOKEN in your .env file!")

    auth = Auth.Token(my_github_token)
    g = Github(auth=auth)

    try:
        repo = g.get_repo(repo_name)
        logger.info(f"开始爬取仓库: {repo_name}")
        logger.info(f"仓库描述: {repo.description}")
    except GithubException as e:
        logger.error(f"无法访问仓库 {repo_name}: {e}")
        return None, None

    # 加载保存的进度
    last_issue_number = load_progress(repo_name)
    logger.info(f"从进度 {last_issue_number} 开始处理。")

    issues_data = []

    # 统计数据
    stats = {
        'total_issues': 0,
        'total_prs': 0,
        'pr_with_images': 0,
        'pr_with_images_b': 0,
        'all_labels': defaultdict(int),
        'issues_with_images': 0,
        'issues_with_images_b': 0,
        'total_images': 0,
        'total_images_b': 0
    }


    # 创建一个新的工作簿和工作表
    wb = Workbook()
    ws = wb.active

    # 创建issue_results目录
    if not os.path.exists('../issue_results'):
        os.makedirs('../issue_results')

    # 设置工作表的表头（添加图片相关字段）
    ws.append([
        'id', 'number', 'html_url', 'type', 'labels', 'created_date', 'updated_date', 'resolved_date', 'title', 'body',
        'state', 'comments', 'state_reason', 'repository_url', 'labels_url', 'comments_url', 'events_url',
        'user_login', 'user_url', 'assignees', 'milestone_title', 'milestone_description', 'pull_request_url',
        'body_image_count', 'comment_image_count', 'total_image_count'
    ])

    # 保存工作表
    start_time = time.time()
    start_time_str = time.strftime("%Y-%m-%d_%H-%M-%S")
    repo_name_temp = repo_name.replace('/', '_')
    excel_path = f'../issue_results/{repo_name_temp}_issues_with_analysis_{start_time_str}.xlsx'
    wb.save(excel_path)

    row_num = len(ws['A']) + 1

    try:
        for issue in repo.get_issues(state="all", labels=["Type: Feature request"]):
            if last_issue_number and issue.number >= last_issue_number:
                continue  # 跳过已处理的 issues

            labels = [label.name for label in issue.labels]
            assignees = [assignee.login for assignee in issue.assignees]
            milestone_title = issue.milestone.title if issue.milestone else None
            milestone_description = issue.milestone.description if issue.milestone else None
            pull_request_url = issue.pull_request.html_url if issue.pull_request else None

            # 判断是issue还是PR
            is_pr = issue.pull_request is not None

            stats['total_issues'] += 1

            if is_pr:
                stats['total_prs'] += 1


            # 收集标签统计
            for label in labels:
                stats['all_labels'][label] += 1

            # 分析正文中的图片
            body_images = analyze_images_in_content(issue.body or "")
            body_image_count = (len(body_images['markdown_images']) +
                                len(body_images['html_images']))

            # 分析评论中的图片（简化版，只统计数量）
            comment_image_count = 0
            try:
                for comment in issue.get_comments():
                    comment_images = analyze_images_in_content(comment.body or "")
                    comment_image_count += (len(comment_images['markdown_images']) +
                                            len(comment_images['html_images']))
            except RateLimitExceededException:
                # 如果在获取评论时遇到速率限制，跳过这个issue的评论分析
                logger.warning(f"  获取issue #{issue.number}的评论时遇到速率限制，跳过评论分析")
            except Exception as e:
                logger.warning(f"  获取issue #{issue.number}评论失败 - {e}")

            total_images_in_issue = body_image_count + comment_image_count

            if body_image_count > 0:
                stats['issues_with_images_b'] += 1
                stats['total_images_b'] += total_images_in_issue
                if is_pr:
                    stats['pr_with_images_b'] += 1

            # 更新统计
            if total_images_in_issue > 0:
                stats['issues_with_images'] += 1
                stats['total_images'] += total_images_in_issue
                if is_pr:
                    stats['pr_with_images'] += 1



            issue_data = {
                'id': issue.id,
                'number': issue.number,
                'html_url': issue.html_url,
                'type': 'issue' if not issue.pull_request else 'pull_request',
                'labels': ','.join(labels),
                'created_date': issue.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                'updated_date': issue.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
                'resolved_date': issue.closed_at.strftime("%Y-%m-%d %H:%M:%S") if issue.closed_at else None,
                'title': issue.title,
                'body': issue.body,
                'state': issue.state,
                'comments': issue.comments,
                'state_reason': issue.state_reason,
                'repository_url': issue.repository.url,
                'labels_url': issue.labels_url,
                'comments_url': issue.comments_url,
                'events_url': issue.events_url,
                'user_login': issue.user.login,
                'user_url': issue.user.html_url,
                'assignees': ','.join(assignees),
                'milestone_title': milestone_title,
                'milestone_description': milestone_description,
                'pull_request_url': pull_request_url,
                'body_image_count': body_image_count,
                'comment_image_count': comment_image_count,
                'total_image_count': total_images_in_issue
            }
            issues_data.append(issue_data)

            logger.info(f'当前已获取 issue #{issue.number} (images: {total_images_in_issue})')

            if len(issues_data) % 100 == 0:  # 减少批量大小，更频繁保存
                issues_df = pd.DataFrame(issues_data)
                for _, row in issues_df.iterrows():
                    if write_row_to_excel(ws, row, row_num):
                        row_num += 1

                wb.save(excel_path)
                logger.info(f'已将 {row_num - 2} 个issues保存到了Excel')
                save_progress(repo_name, issue.number)  # 保存当前最小编号进度

                # 保存中间统计结果
                save_intermediate_stats(repo_name, stats, len(issues_data))
                issues_data = []

        # 保存剩余数据
        if issues_data:
            issues_df = pd.DataFrame(issues_data)
            for _, row in issues_df.iterrows():
                if write_row_to_excel(ws, row, row_num):
                    row_num += 1

        wb.save(excel_path)
        save_progress(repo_name, 0)  # 清空进度，表示完成
        logger.info(f'导出成功，共有 {row_num - 2} 个issues')

        # 保存最终统计结果
        save_final_stats(repo_name, stats)

        return excel_path, stats

    except RateLimitExceededException:
        reset_time = g.get_rate_limit().core.reset.timestamp()
        sleep_time = reset_time - time.time() + 1
        logger.warning(f"速率限制已达到，暂停 {sleep_time} 秒。")
        time.sleep(sleep_time)
        return get_issue_to_excel_with_analysis(repo_name)  # 在同一运行中重试

    except GithubException as e:
        logger.error(f"获取仓库 {repo_name} 的数据时出错: {e}")
        # 保存当前进度和数据
        if issues_data:
            save_intermediate_stats(repo_name, stats, len(issues_data))
        return excel_path if 'excel_path' in locals() else None, stats


def save_intermediate_stats(repo_name, stats, processed_count):
    """保存中间统计结果"""
    repo_name_clean = repo_name.replace('/', '_')
    stats_file = f'../issue_results/{repo_name_clean}_stats_progress_{processed_count}.json'

    # 转换defaultdict为普通dict以便JSON序列化
    stats_copy = dict(stats)
    stats_copy['all_labels'] = dict(stats_copy['all_labels'])

    with open(stats_file, 'w', encoding='utf-8') as f:
        json.dump(stats_copy, f, ensure_ascii=False, indent=2)


def save_final_stats(repo_name, stats):
    """保存最终统计结果"""
    repo_name_clean = repo_name.replace('/', '_')
    stats_file = f'../issue_results/{repo_name_clean}_final_statistics.json'

    # 转换defaultdict为普通dict以便JSON序列化
    stats_copy = dict(stats)
    stats_copy['all_labels'] = dict(stats_copy['all_labels'])

    with open(stats_file, 'w', encoding='utf-8') as f:
        json.dump(stats_copy, f, ensure_ascii=False, indent=2)

    logger.info(f"统计数据已保存到: {stats_file}")


def print_statistics(stats):
    """打印统计结果"""
    print(f"\n{'=' * 50}")
    print(f"  仓库 Issues 统计分析结果")
    print(f"{'=' * 50}")

    print(f"\n=== 基本统计 ===")
    print(f"总Issues数: {stats['total_issues']}")
    print(f"总PRs数: {stats['total_prs']}")
    print(f"包含图片的issues: {stats['issues_with_images']}")
    print(f"包含图片的PRs: {stats['pr_with_images']}")
    print(f"包含图片的issues（增强型）: {stats['issues_with_images_b']}")
    print(f"包含图片的PRs（增强型）: {stats['pr_with_images_b']}")
    print(f"总图片数: {stats['total_images']}")
    print(f"总图片数（增强型）: {stats['total_images_b']}")


    if stats['total_issues'] + stats['total_prs'] > 0:
        image_rate = (stats['issues_with_images'] / (stats['total_issues'] + stats['total_prs'])) * 100
        print(f"包含图片的比例: {image_rate:.1f}%")





    print(f"\n{'=' * 50}")


if __name__ == "__main__":
    repo_name = "AntennaPod1/AntennaPod1"

    print(f"开始使用issue_2逻辑爬取和分析仓库: {repo_name}")
    print("程序会自动处理速率限制并从上次中断的地方继续执行。")
    print("这可能需要一些时间，请耐心等待...")

    excel_path, stats = get_issue_to_excel_with_analysis(repo_name)

    if stats and excel_path:
        print(f"\n爬取完成！Excel文件已保存到: {excel_path}")
        print_statistics(stats)
    else:
        print("分析失败或被中断！请重新运行程序继续处理。")

